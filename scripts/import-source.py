"""Rebuild the app seed from the authoritative saving workbook.

Source: finance_project___update saving_20260709 (2).xlsx, sheet "Project".
That sheet — not the Jira export — is the system of record for saving hours.
Jira keys are carried where present so the two can be reconciled later.
"""
import openpyxl, sys, io, json, re, os
from collections import defaultdict
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SRC = r"c:\Users\th40184213\Downloads\FA-tech-team-objective-2026\finance_project___update saving_20260709 (2).xlsx"
OUT = r"c:\Users\th40184213\Downloads\FA-tech-team-objective-2026\team-kpi-planning\src\data\seed.json"

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["Project"]

# The six scorecard owners, plus IT as an assignable delivery owner. IT carries
# no scorecard of its own - it is a partner team - so its hours fall to the
# team lead, who is accountable for the team's overall KPI.
CORE = {
    # aggregatesTeam: the lead is measured on everything the team delivers -
    # their own projects, IT-owned and unassigned work, and every other member's
    # credited hours - because they carry the team's overall KPI.
    "gun":   {"id": "gun",   "name": "Wisarut Gunjarueg", "nick": "Gun",      "role": "Team Lead", "band": "lead",    "scorecard": True, "aggregatesTeam": True},
    "james": {"id": "james", "name": "Pipat Singhasiri",  "nick": "James",    "role": "Senior",    "band": "senior",  "scorecard": True},
    "pphen": {"id": "pphen", "name": "Chanphen Manu",     "nick": "P'Phen",   "role": "Senior",    "band": "senior",  "scorecard": True},
    "kade":  {"id": "kade",  "name": "Jarinya Phosri",    "nick": "Kade",     "role": "Senior",    "band": "senior",  "scorecard": True},
    "tha":   {"id": "tha",   "name": "Thapanee",          "nick": "Thapanee", "role": "Analyst",   "band": "analyst", "scorecard": True},
    "pol":   {"id": "pol",   "name": "Pol",               "nick": "Pol",      "role": "Analyst",   "band": "analyst", "scorecard": True},
    "it":    {"id": "it",    "name": "IT",                "nick": "IT",       "role": "Partner team", "band": "senior", "scorecard": False},
}

# Assignee string (lowercased) -> owner id. Anything unmatched is left without a
# PIC; those hours fall to the team lead via the fallback rule in the app.
ASSIGNEE = {
    "wisarut gunjarueg": "gun", "wisarut": "gun",
    "pipat.singhasiri": "james",
    "chanphen": "pphen",
    "jarinya phosri": "kade",
    "thapanee": "tha", "thanpanee": "tha",
    "it": "it",
}


def fix_mojibake(s):
    """Repair text that was UTF-8 encoded then decoded as Latin-1.

    The source workbook carries a few cells like "â€” Actual & Forecast", which
    is an em dash whose three UTF-8 bytes were each read as a separate Latin-1
    character. Round-tripping back through Latin-1 recovers the original.
    """
    if not s or not any(m in s for m in ("â", "Ã", "Â")):
        return s
    # cp1252 first: the mojibake usually contains characters such as € and "
    # that Latin-1 cannot encode, which is what makes a naive round-trip fail.
    for codec in ("cp1252", "latin-1"):
        try:
            repaired = s.encode(codec).decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            continue
        # only accept a repair that actually removed the tell-tale sequences
        if repaired.count("â") < s.count("â") or repaired.count("Ã") < s.count("Ã"):
            return repaired
    return s


def clean(v):
    if v is None:
        return None
    s = fix_mojibake(str(v).strip())
    return s or None


def norm_date(v):
    """The sheet has typo years (1926 for 2026). Repair rather than drop."""
    s = clean(v)
    if not s:
        return None
    s = s.split(" ")[0]
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if not m:
        return None
    y = int(m.group(1))
    if y < 2000:
        y += 100  # 1926 -> 2026
    return f"{y:04d}-{m.group(2)}-{m.group(3)}"


def fnum(v):
    s = clean(v)
    if s is None or s.upper() == "TBC":
        return None
    try:
        return float(str(s).replace(",", ""))
    except ValueError:
        return None


STATUS_MAP = {"done": "Done", "in progress": "In Progress", "not start": "Not Start"}
COMMIT_BY_STATUS = {"Done": "commit", "In Progress": "commit", "Not Start": "stretch"}


def classify(detail, program, subteam):
    s = f"{detail or ''} {program or ''}".lower()
    if any(k in s for k in ("datacube", "datawarehouse", "data warehouse", "data library",
                            "databricks", "master file centralization", "margin database")):
        return "datawarehouse"
    if any(k in s for k in ("ocr", "chatbot", "rpa", "bot", "automail", "auto email",
                            "agent", "genie", "ai ")):
        return "ai_automation"
    if any(k in s for k in ("dashboard", "pbi", "portal", "report", "p&l", "pnl",
                            "summary", "hub", "webapp", "app")):
        return "efficiency"
    if any(k in s for k in ("forecast", "margin", "budget", "loss", "shrinkage",
                            "payroll", "opex", "provision", "coupon", "sales")):
        return "financial"
    return "process_automation"


rows = []
seen = defaultdict(int)
for r in range(2, ws.max_row + 1):
    detail = clean(ws.cell(r, 7).value)          # G Detail
    program = clean(ws.cell(r, 4).value)         # D Project
    if not detail and not program:
        continue

    key = clean(ws.cell(r, 1).value)             # A Key
    team = clean(ws.cell(r, 2).value)            # B Team
    subteam = clean(ws.cell(r, 3).value)         # C Sub Team
    saving = fnum(ws.cell(r, 5).value)           # E Saving hrs/mth
    hc = fnum(ws.cell(r, 6).value)               # F HC
    status = STATUS_MAP.get((clean(ws.cell(r, 8).value) or "").lower(), clean(ws.cell(r, 8).value) or "Not Start")
    assignee = clean(ws.cell(r, 9).value)        # I Assignee
    start = norm_date(ws.cell(r, 10).value)
    due = norm_date(ws.cell(r, 11).value)
    remark = clean(ws.cell(r, 12).value)         # L Remark

    pic = ASSIGNEE.get((assignee or "").lower().strip())

    # Stable unique id: a Jira key can legitimately appear on several rows
    # (FNP-1151 is split across three sub-teams), so suffix repeats.
    base = key or f"ROW-{r}"
    seen[base] += 1
    uid = base if seen[base] == 1 else f"{base}#{seen[base]}"

    obj = classify(detail, program, subteam)
    commit = COMMIT_BY_STATUS.get(status, "stretch")
    if saving is None:
        commit = "watch"   # TBC benefit cannot be committed

    rows.append({
        "key": uid,
        "jiraKey": key,
        "summary": detail or program,
        "program": program,
        "team": team,
        "subTeam": subteam,
        "objective": obj,
        "savingHours": saving,          # hours per MONTH (per the source column header)
        "hc": hc,
        "savingEstimated": saving is None,
        "manday": 0,
        "mandayEstimated": True,
        "status": status,
        "srcStatus": None,
        "start": start,
        "due": due,
        "assignee": assignee,
        "pic": pic,
        "contributors": [{"person": pic, "roles": ["dev"]}] if pic else [],
        "partners": [] if pic else ([{"person": assignee, "roles": ["dev"]}] if assignee else []),
        "deleted": False,
        "commitLevel": commit,
        "notes": remark or "",
    })

seed = {
    "meta": {
        "source": "finance_project___update saving_20260709 (2).xlsx / sheet Project",
        "generated": "2026-08-07",
        "savingBasis": "monthly",
        "note": "Saving hours are PER MONTH (source column: 'Saving hrs/mth'). Mandays are not in the source and start at 0 - enter them in the app.",
    },
    "people": list(CORE.values()),
    "projects": rows,
}
os.makedirs(os.path.dirname(OUT), exist_ok=True)
json.dump(seed, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

# ---------------- report ----------------
tot = sum(p["savingHours"] or 0 for p in rows)
print(f"rows: {len(rows)}  -> {OUT}")
print(f"TOTAL saving hrs/mth: {tot:,.1f}   (= {tot*12:,.0f} hrs/year)")
print(f"total HC: {sum(p['hc'] or 0 for p in rows):,.1f}")
print(f"rows with a Jira key: {sum(1 for p in rows if p['jiraKey'])}, without: {sum(1 for p in rows if not p['jiraKey'])}")
print(f"TBC saving: {sum(1 for p in rows if p['savingHours'] is None)}")

print("\nby Team:")
for t in sorted({p["team"] for p in rows if p["team"]}):
    sub = [p for p in rows if p["team"] == t]
    print(f"  {t:<12} rows={len(sub):<3} hrs/mth={sum(p['savingHours'] or 0 for p in sub):>8,.1f}  HC={sum(p['hc'] or 0 for p in sub):>5,.1f}")

print("\nby status:")
for s in ("Done", "In Progress", "Not Start"):
    sub = [p for p in rows if p["status"] == s]
    print(f"  {s:<12} rows={len(sub):<3} hrs/mth={sum(p['savingHours'] or 0 for p in sub):>8,.1f}  HC={sum(p['hc'] or 0 for p in sub):>5,.1f}")

print("\nby PIC:")
for pid in list(CORE) + [None]:
    sub = [p for p in rows if p["pic"] == pid]
    if not sub:
        continue
    nick = CORE[pid]["nick"] if pid else "(no core PIC)"
    print(f"  {nick:<14} rows={len(sub):<3} hrs/mth={sum(p['savingHours'] or 0 for p in sub):>8,.1f}")

print("\ntop 10 by saving hrs/mth:")
for p in sorted(rows, key=lambda x: -(x["savingHours"] or 0))[:10]:
    print(f"  {p['key']:<12} {p['savingHours'] or 0:>8,.1f}  {p['team']:<11} {p['summary'][:44]}")
